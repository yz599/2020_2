"""
输出
/Users/Yi/我的python教程/pandas办公自动化_备课/data_out/2019全平台销售总表.xlsx
"""
import pandas as pd
import numpy as np
import os
import sys
from time import time

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

pd.options.display.max_columns = 888
pd.options.display.max_rows = 888

out_xlsx_title_1 = ['货号', '商品名称', '图片', '尺码', '日期', '售卖价',
                    '实收金额', '销售量_含拒退', '销售量_不含拒退',
                    '平台',
                    '品牌', '成本', '季节', '商品年份']


# 函数运行时间装饰器
def use_time(func):
    def wrapper(*args, **kwargs):
        start = time()
        result = func(*args, **kwargs)
        use = round(time() - start, 3)
        print('%s()用时: %s秒' % (func.__name__, use))
        return result

    return wrapper


# 向 excel 插入图片
def save_image_to_excel(file_path, picture_path, columns, sheet_index=0):
    """
    把图片插入到xlsx
    :param file_path: 文件路径
    :param picture_path: 图片路径
    :param columns: 2D_list, [[0, B], [2, D]]表示: 依据第0列, 插入到到B列; 依据第2列, 插入到到D列
    :param sheet_index: 选择sheet, 从0开现
    :return: None
    """
    print('插入图片...')
    # img_path = '//10.1.4.184/g/data/picture/X89113075049.png'
    # wb = load_workbook(r'G:\yi\yi_code\hxfs_data_y\my_module\1.xlsx')
    df = pd.read_excel(file_path, sheet_index=sheet_index)
    wb = load_workbook(file_path)
    ws = wb.worksheets[sheet_index]
    for depend_col, insert_col in columns:
        depend_list = df.iloc[:, depend_col].tolist()
        for idx, pic_name in enumerate(depend_list, start=1):
            try:
                img = Image(os.path.join(picture_path, pic_name + '.jpg'))
            except:
                try:
                    img = Image(os.path.join(picture_path, pic_name + '.png'))
                except:
                    continue
            anchor = insert_col + str(idx + 1)
            ws.add_image(img, anchor=anchor)
    wb.save(file_path)
    print('插入图片完成')


def concat_files(folder):
    """
    读取文件夹内的所有表格, 合并成一个表
    :param folder: 文件夹路径
    :return: DataFrame
    """
    df_list = []
    for fn in os.listdir(folder):
        ffn = os.path.join(folder, fn)
        if '.xlsx' not in ffn:
            continue
        # print(ffn)
        df_temp = pd.read_excel(ffn)
        df_list.append(df_temp)
    df = pd.concat(df_list)
    return df


# ['货号', '商品名称', '尺码', '日期', '售卖价', '实收金额', '销售量_含拒退', '销售量_不含拒退']
def read_sales_b():
    """
    ['款号颜色代码', '商品名称', '尺码', '日期', '售卖价', '活动价', '销售量_含拒退', '销售量_不含拒退']
    """
    df_b = concat_files(
        './PlatformB'
    )
    df_b['实收金额'] = np.where(df_b['活动价'] == 0, df_b['售卖价'], df_b['活动价'])
    df_b = df_b.rename(columns={'款号颜色代码': '货号'})
    df_b = df_b.drop(columns=['活动价'])
    return df_b


@use_time
def out_sales_summary_tb():
    print('输出"2019全平台销售总表.xlsx"...')
    df_a = concat_files('./PlatformA')
    df_a['平台'] = '平台A'
    df_b = read_sales_b()
    df_b['平台'] = '平台B'
    df_all = df_a.append(df_b)
    # ['售卖价', '商品名称', '实收金额', '尺码', '平台',
    # '日期', '货号', '销售量_不含拒退', '销售量_含拒退']
    # print(df_all.head())
    df_base = pd.read_excel(
        '商品基础信息表.xlsx',
        usecols=['货号', '品牌', '成本', '季节', '商品年份']
    )
    df_all = pd.merge(df_all, df_base, on='货号', how='left',
                      validate='many_to_one')
    df_all['日期'] = df_all['日期'].dt.date
    # print(df_all.head())

    df_all.to_excel('./data_out/2019全平台销售总表.xlsx', index=None)
    # df_all.to_pickle('./data_out/2019全平台销售总表.pkl')
    print('输出"2019全平台销售总表.xlsx"完成', end='\n**********\n')


# 压缩图片
@use_time
def compress_image(src_path, dst_path, w=80, h=80):
    print('压缩图片...')
    src_fn_set = set(os.listdir(src_path))
    dst_fn_set = set(os.listdir(dst_path))
    fn_set = src_fn_set - dst_fn_set
    if len(fn_set) == 0:
        print('无可更新图片', end='\n**********\n')
        return

    from PIL import Image
    for fn in fn_set:
        src_ffn = os.path.join(src_path, fn)
        dst_ffn = os.path.join(dst_path, fn)

        larger_img = Image.open(src_ffn)
        small_img = larger_img.resize((w, h), resample=Image.ANTIALIAS)
        small_img.save(dst_ffn)

        print(src_ffn)
        print(dst_ffn)
    print('压缩图片完成', end='\n**********\n')


"""
输出
./data_out/2019销售汇总表.xlsx
"""
out_xlsx_title = ['货号', '商品名称', '图片', '总销量', '总销售额',
                  '总成本', '总利润']


@use_time
def out_goods_id_total():
    print('输出"2019销售汇总表.xlsx"...')
    # df = pd.read_pickle('./data_out/2019全平台销售总表.pkl')
    df = pd.read_excel('./data_out/2019全平台销售总表.xlsx')
    # print(df.head())
    df = df[['货号', '商品名称', '实收金额', '销售量_不含拒退', '成本']]
    df = df.rename(columns={'销售量_不含拒退': '总销量'})
    df['总销售额'] = df['实收金额'] * df['总销量']
    df = pd.pivot_table(
        df,
        index=['货号', '商品名称', '成本'],
        values=['总销量', '总销售额'],
        aggfunc=sum)
    df = df.reset_index()
    df['总成本'] = df['总销量'] * df['成本']
    df['总利润'] = df['总销售额'] - df['总成本']

    total_series = df[['总销量', '总销售额', '总成本', '总利润']].sum(axis=0).reindex(out_xlsx_title)
    total_series['货号'] = '合计'
    # print(total_series)
    # return
    df = df.append(total_series, ignore_index=True)

    df = df.reindex(columns=out_xlsx_title)
    # print(df)
    ffn = './data_out/2019销售汇总表.xlsx'
    df.to_excel(ffn, index=None)
    # 插入图片
    picture_path = './压缩图'
    save_image_to_excel(ffn, picture_path, [[0, 'C']])
    print('输出"2019销售汇总表.xlsx"完成', end='\n**********\n')


if __name__ == '__main__':
    # 1: 压缩图片
    compress_image('./原图', './压缩图')

    # 2: 输出"2019全平台销售总表.xlsx"
    out_sales_summary_tb()

    # 3: 输出"2019销售汇总表.xlsx"
    out_goods_id_total()
